#!/usr/bin/env python3

from __future__ import annotations

from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


REPO_ROOT = Path(__file__).resolve().parents[2]
RAW_EXPORT_PATH = REPO_ROOT / "analysis" / "validity data" / "data.xlsx"

DATA_CLEAN_DIR = REPO_ROOT / "analysis" / "data_clean"
OUTPUTS_DIR = REPO_ROOT / "analysis" / "outputs"
TABLES_DIR = REPO_ROOT / "tables"
STYLED_PREVIEW_DIR = TABLES_DIR / "styled_preview"
FIGURES_DIR = REPO_ROOT / "figures"

VALIDATION_EXPERTS_PATH = DATA_CLEAN_DIR / "validation_experts_clean.csv"
ITEMS_CLEAN_PATH = DATA_CLEAN_DIR / "rater_items_expert_review_clean.csv"
SEEDS_CLEAN_PATH = DATA_CLEAN_DIR / "seeds_expert_review_clean.csv"

PANEL_SUMMARY_PATH = OUTPUTS_DIR / "expert_review_panel_summary.csv"
PANEL_BREAKDOWN_PATH = OUTPUTS_DIR / "expert_review_panel_breakdown.csv"
ITEM_SUMMARY_PATH = OUTPUTS_DIR / "expert_review_item_summary.csv"
SEED_SUMMARY_PATH = OUTPUTS_DIR / "expert_review_seed_summary.csv"
COMMENT_LOG_PATH = OUTPUTS_DIR / "expert_review_comment_log.csv"

TABLE_ITEMS_PATH = TABLES_DIR / "table_s4_expert_review_items.csv"
TABLE_SEEDS_PATH = TABLES_DIR / "table_s5_expert_review_seeds.csv"

FIGURE_ITEMS_PATH = FIGURES_DIR / "figure_s3_expert_review_items_heatmap.png"
FIGURE_SEEDS_PATH = FIGURES_DIR / "figure_s4_expert_review_seeds_heatmap.png"

HTML_ITEMS_PATH = STYLED_PREVIEW_DIR / "table_s4_expert_review_items.html"
HTML_SEEDS_PATH = STYLED_PREVIEW_DIR / "table_s5_expert_review_seeds.html"
HTML_PREVIEW_PATH = STYLED_PREVIEW_DIR / "expert_review_preview.html"

PROFESSION_MAP = {
    "1": "psychológ / psychologička",
    "2": "psychiater / psychiatrička",
    "3": "lekár v príprave v psychiatrii",
    "4": "psychoterapeut / psychoterapeutka",
    "5": "iný odborník / iná odborníčka v duševnom zdraví",
    "6": "iné",
}

SPECIALIZATION_MAP = {
    "1": "klinická psychológia",
    "2": "poradenská psychológia",
    "3": "psychiatrická ambulancia",
    "4": "inpatient / lôžková psychiatria",
    "5": "psychoterapia",
    "6": "výskum",
    "7": "iné",
}

EXPERIENCE_MAP_4 = {
    "1": "žiadna",
    "2": "nízka / skôr menšia / občasná",
    "3": "stredná / pravidelná",
    "4": "vysoká / veľmi častá / výrazná",
}

TOOL_ITEMS = [
    ("G1", "G", "Klinická vierohodnosť rozhovoru"),
    ("G2", "G", "Prirodzenosť jazyka a štýlu odpovedí"),
    ("G3", "G", "Vnútorná konzistentnosť rozhovoru"),
    ("G4", "G", "Súlad rozhovoru s obrazom depresívnej symptomatiky"),
    ("G5", "G", "Použiteľnosť rozhovoru na tréningové / výučbové účely"),
    ("S1", "S", "Odhad celkovej závažnosti depresívnej symptomatiky"),
    ("S2", "S", "Odhad funkčného dopadu na bežné fungovanie"),
    ("R1", "R", "Prítomnosť vnútorných kontradikcií v rozhovore"),
    ("R2", "R", "Klišé alebo šablónovité odpovede"),
    ("R3", "R", "Nesúlad medzi kontextom a symptomatikou"),
    ("R4", "R", "Podozrenie na inú primárnu psychopatológiu"),
    ("R5", "R", "Neprimeraná dramatizácia alebo neprirodzená expresivita"),
]

SEED_TITLES = {
    1: "Mladá žena s pracovno-študijným preťažením",
    2: "Mladý muž so smennou prácou",
    3: "Žena v strednom veku s caregiving záťažou",
    4: "Muž s vysokou pracovnou záťažou",
    5: "Rozvedená žena s finančným tlakom",
    6: "Muž s dlhou pracovnou záťažou",
    7: "Staršia vdova so stratou rytmu",
    8: "Starší muž po strate rutiny",
    9: "Výkonovo zaťažená žena s rumináciami",
    10: "Muž s rutinou, stagnáciou a pasívnym A9",
    11: "Žena s nízkym príjmom a sociálnym stiahnutím",
    12: "Muž s vysokou zodpovednosťou a utlmeným prežívaním",
}

SEED_CRITERIA = [
    ("ambulatory_profile", "Zodpovedá bežnému ambulantnému profilu"),
    ("context_realism", "Realistický demografický a sociálny kontext"),
    ("symptom_plausibility", "Klinická vierohodnosť symptómov"),
    ("clarity", "Jasnosť a zrozumiteľnosť"),
    ("information_sufficiency", "Dostatok informácií pre generovanie rozhovoru"),
    ("anti_stereotypy", "Nepôsobí stereotypne alebo karikatúrne"),
    ("research_usefulness", "Vhodnosť na použitie vo výskume"),
]

SEVERITY_LABELS = {
    1: "skôr mierna",
    2: "skôr stredná",
    3: "skôr stredne ťažká",
    4: "neviem posúdiť",
}

TARGET_SEVERITY = {
    1: 2,
    2: 2,
    3: 3,
    4: 2,
    5: 2,
    6: 2,
    7: 2,
    8: 1,
    9: 2,
    10: 2,
    11: 2,
    12: 2,
}


def ensure_dirs() -> None:
    for path in (DATA_CLEAN_DIR, OUTPUTS_DIR, TABLES_DIR, STYLED_PREVIEW_DIR, FIGURES_DIR):
        path.mkdir(parents=True, exist_ok=True)


def read_named_columns_from_xlsx(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [worksheet.cell(1, idx).value for idx in range(1, worksheet.max_column + 1)]
    named_positions = [idx for idx, header in enumerate(headers) if header not in (None, "")]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    frame = pd.DataFrame(
        [[row[idx] for idx in named_positions] for row in rows],
        columns=[headers[idx] for idx in named_positions],
    )
    return frame


def text_or_blank(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def recode_label(raw_value: object, mapping: dict[str, str], fallback: str = "") -> str:
    key = text_or_blank(raw_value)
    if not key:
        return fallback
    return mapping.get(key, fallback or key)


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_candidates = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Helvetica.ttc",
    ]
    for candidate in font_candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = text.split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if draw.textlength(candidate, font=font) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def lerp_color(start: tuple[int, int, int], end: tuple[int, int, int], ratio: float) -> tuple[int, int, int]:
    return tuple(int(start[idx] + (end[idx] - start[idx]) * ratio) for idx in range(3))


def color_for_score(score: float, low: float = 1.0, high: float = 4.0) -> tuple[int, int, int]:
    start = (245, 228, 215)
    end = (45, 106, 79)
    ratio = 0.0 if high == low else max(0.0, min(1.0, (score - low) / (high - low)))
    return lerp_color(start, end, ratio)


def draw_heatmap(
    *,
    data: pd.DataFrame,
    row_label_col: str,
    value_cols: list[str],
    display_cols: list[str],
    title: str,
    subtitle: str,
    output_path: Path,
) -> None:
    title_font = load_font(28)
    subtitle_font = load_font(18)
    header_font = load_font(17)
    label_font = load_font(17)
    cell_font = load_font(16)
    note_font = load_font(15)

    cell_width = 130
    cell_height = 42
    label_width = 420
    top_margin = 120
    left_margin = 28
    right_margin = 24
    bottom_margin = 80

    image_width = left_margin + label_width + len(value_cols) * cell_width + right_margin
    image_height = top_margin + (len(data) + 1) * cell_height + bottom_margin
    image = Image.new("RGB", (image_width, image_height), color=(252, 249, 244))
    draw = ImageDraw.Draw(image)

    draw.text((left_margin, 20), title, fill=(32, 37, 41), font=title_font)
    draw.text((left_margin, 58), subtitle, fill=(86, 92, 100), font=subtitle_font)

    header_y = top_margin
    draw.rectangle(
        [left_margin, header_y, left_margin + label_width + len(value_cols) * cell_width, header_y + cell_height],
        fill=(236, 232, 225),
        outline=(185, 179, 168),
    )
    draw.text((left_margin + 12, header_y + 11), "Položka", fill=(32, 37, 41), font=header_font)
    for idx, header in enumerate(display_cols):
        x0 = left_margin + label_width + idx * cell_width
        draw.rectangle([x0, header_y, x0 + cell_width, header_y + cell_height], outline=(185, 179, 168))
        header_lines = wrap_text(draw, header, header_font, cell_width - 16)
        for line_idx, line in enumerate(header_lines[:2]):
            draw.text((x0 + 10, header_y + 6 + line_idx * 16), line, fill=(32, 37, 41), font=header_font)

    for row_idx, (_, row) in enumerate(data.iterrows(), start=1):
        y0 = top_margin + row_idx * cell_height
        y1 = y0 + cell_height
        draw.rectangle(
            [left_margin, y0, left_margin + label_width, y1],
            fill=(248, 245, 239) if row_idx % 2 else (244, 240, 233),
            outline=(220, 214, 205),
        )
        label_lines = wrap_text(draw, str(row[row_label_col]), label_font, label_width - 16)
        for line_idx, line in enumerate(label_lines[:2]):
            draw.text((left_margin + 12, y0 + 8 + line_idx * 15), line, fill=(42, 46, 52), font=label_font)
        for col_idx, value_col in enumerate(value_cols):
            x0 = left_margin + label_width + col_idx * cell_width
            score = float(row[value_col])
            fill = color_for_score(score)
            draw.rectangle([x0, y0, x0 + cell_width, y1], fill=fill, outline=(220, 214, 205))
            text_fill = (248, 248, 248) if score >= 3.2 else (31, 37, 41)
            text = f"{score:.2f}"
            text_width = draw.textlength(text, font=cell_font)
            draw.text((x0 + (cell_width - text_width) / 2, y0 + 12), text, fill=text_fill, font=cell_font)

    legend_y = image_height - 48
    draw.text((left_margin, legend_y), "Škála: 1 = slabé hodnotenie, 4 = silná podpora.", fill=(86, 92, 100), font=note_font)

    image.save(output_path)


def write_html_table(df: pd.DataFrame, title: str, note: str, output_path: Path) -> None:
    table_html = df.to_html(index=False, border=0, classes="thesis-table")
    html = f"""<!doctype html>
<html lang="sk">
<head>
  <meta charset="utf-8">
  <title>{title}</title>
  <style>
    body {{
      margin: 0;
      background: #fbf7f2;
      color: #202529;
      font-family: Georgia, 'Times New Roman', serif;
    }}
    .wrap {{
      max-width: 1240px;
      margin: 0 auto;
      padding: 40px 48px;
    }}
    .caption {{
      font-size: 28px;
      font-weight: 700;
      margin-bottom: 8px;
    }}
    .note {{
      font-size: 16px;
      color: #5e646b;
      margin-bottom: 20px;
      line-height: 1.45;
    }}
    table.thesis-table {{
      width: 100%;
      border-collapse: collapse;
      background: #fffdfa;
      font-size: 14px;
      line-height: 1.35;
    }}
    table.thesis-table thead th {{
      text-align: left;
      padding: 10px 12px;
      border-top: 2px solid #525a61;
      border-bottom: 1px solid #b2b8bf;
      background: #efeae2;
    }}
    table.thesis-table tbody td {{
      padding: 9px 12px;
      border-bottom: 1px solid #e1ddd6;
      vertical-align: top;
    }}
    table.thesis-table tbody tr:nth-child(even) {{
      background: #faf6f0;
    }}
    img {{
      width: 100%;
      height: auto;
      border: 1px solid #d9d1c5;
      background: #fffdfa;
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="caption">{title}</div>
    <div class="note">{note}</div>
    {table_html}
  </div>
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


def build_validation_experts(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for idx, record in df.iterrows():
        profession_label = recode_label(record["profession"], PROFESSION_MAP)
        profession_other = text_or_blank(record["profession_other"])
        if profession_label == "iné" and profession_other:
            profession_label = profession_other
        rows.append(
            {
                "expert_id": f"E{idx + 1:02d}",
                "source_participant_id": text_or_blank(record["participant"]),
                "profession": profession_label,
                "specialization": recode_label(record["specialization"], SPECIALIZATION_MAP),
                "years_practice": int(record["years_practice"]),
                "depression_experience": recode_label(record["depression_experience"], EXPERIENCE_MAP_4),
                "interview_experience": recode_label(record["interview_experience"], EXPERIENCE_MAP_4),
                "teaching_or_supervision_experience": recode_label(record["teaching_experience"], EXPERIENCE_MAP_4),
                "ai_experience": recode_label(record["ai_experience"], EXPERIENCE_MAP_4),
                "tool_comment": text_or_blank(record["tool_comment"]),
                "seed_overall_comment": text_or_blank(record["seed_overall"]),
            }
        )
    return pd.DataFrame(rows)


def build_item_review(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for expert_idx, record in df.iterrows():
        expert_id = f"E{expert_idx + 1:02d}"
        for item_idx, (item_code, block, label) in enumerate(TOOL_ITEMS, start=1):
            relevance = int(record[f"tool_rel_{item_idx}"])
            clarity = int(record[f"tool_clarity_{item_idx}"])
            need = int(record[f"tool_need_{item_idx}"])
            rows.append(
                {
                    "expert_id": expert_id,
                    "item_id": item_idx,
                    "item_code": item_code,
                    "block": block,
                    "item_label": label,
                    "relevance_1_4": relevance,
                    "clarity_1_4": clarity,
                    "need_1_4": need,
                    "item_mean_1_4": round((relevance + clarity + need) / 3, 3),
                }
            )
    return pd.DataFrame(rows)


def build_seed_review(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for expert_idx, record in df.iterrows():
        expert_id = f"E{expert_idx + 1:02d}"
        for seed_id, title in SEED_TITLES.items():
            criterion_scores = {
                key: int(record[f"seed{seed_id:02d}_eval_{criterion_idx}"])
                for criterion_idx, (key, _) in enumerate(SEED_CRITERIA, start=1)
            }
            severity_code = int(record[f"seed{seed_id:02d}_severity"])
            rows.append(
                {
                    "expert_id": expert_id,
                    "seed_id": seed_id,
                    "seed_title": title,
                    **{f"{key}_1_4": value for key, value in criterion_scores.items()},
                    "overall_mean_1_4": round(sum(criterion_scores.values()) / len(criterion_scores), 3),
                    "target_severity_code": TARGET_SEVERITY[seed_id],
                    "target_severity_label": SEVERITY_LABELS[TARGET_SEVERITY[seed_id]],
                    "severity_rating_code": severity_code,
                    "severity_rating_label": SEVERITY_LABELS[severity_code],
                    "severity_matches_target": int(severity_code == TARGET_SEVERITY[seed_id]),
                    "comment": text_or_blank(record[f"seed{seed_id:02d}_comment"]),
                }
            )
    return pd.DataFrame(rows)


def summarize_panel(
    experts_df: pd.DataFrame,
    items_df: pd.DataFrame,
    seeds_df: pd.DataFrame,
    comment_log: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    years = experts_df["years_practice"]
    summary = pd.DataFrame(
        [
            {
                "n_experts": len(experts_df),
                "n_item_review_rows": len(items_df),
                "n_seed_review_rows": len(seeds_df),
                "n_item_dimension_ratings": len(items_df) * 3,
                "n_seed_dimension_ratings": len(seeds_df) * len(SEED_CRITERIA),
                "n_seed_severity_ratings": len(seeds_df),
                "years_practice_median": float(years.median()),
                "years_practice_min": int(years.min()),
                "years_practice_max": int(years.max()),
                "n_with_any_tool_comment": int(experts_df["tool_comment"].str.len().gt(0).sum()),
                "n_with_seed_overall_comment": int(experts_df["seed_overall_comment"].str.len().gt(0).sum()),
                "n_logged_comments_total": len(comment_log),
            }
        ]
    )

    breakdown_rows = []
    for field in [
        "profession",
        "specialization",
        "depression_experience",
        "interview_experience",
        "teaching_or_supervision_experience",
        "ai_experience",
    ]:
        counts = experts_df[field].value_counts().sort_values(ascending=False)
        for label, count in counts.items():
            breakdown_rows.append(
                {
                    "field": field,
                    "label": label,
                    "n": int(count),
                    "share": round(count / len(experts_df), 3),
                }
            )
    return summary, pd.DataFrame(breakdown_rows)


def summarize_items(items_df: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        items_df.groupby(["item_id", "item_code", "block", "item_label"], as_index=False)
        .agg(
            relevance_mean_1_4=("relevance_1_4", "mean"),
            clarity_mean_1_4=("clarity_1_4", "mean"),
            need_mean_1_4=("need_1_4", "mean"),
            relevance_support_3_4=("relevance_1_4", lambda s: (s >= 3).mean()),
            clarity_support_3_4=("clarity_1_4", lambda s: (s >= 3).mean()),
            need_support_3_4=("need_1_4", lambda s: (s >= 3).mean()),
        )
        .sort_values(["item_id"])
    )
    grouped["overall_mean_1_4"] = grouped[
        ["relevance_mean_1_4", "clarity_mean_1_4", "need_mean_1_4"]
    ].mean(axis=1)
    grouped["overall_support_3_4"] = grouped[
        ["relevance_support_3_4", "clarity_support_3_4", "need_support_3_4"]
    ].mean(axis=1)
    grouped["follow_up_flag"] = grouped[
        ["relevance_support_3_4", "clarity_support_3_4", "need_support_3_4"]
    ].min(axis=1) < 0.833
    return grouped.sort_values(["overall_mean_1_4", "overall_support_3_4", "item_id"]).reset_index(drop=True)


def summarize_seeds(seeds_df: pd.DataFrame) -> pd.DataFrame:
    agg_spec = {
        f"{key}_mean_1_4": (f"{key}_1_4", "mean")
        for key, _ in SEED_CRITERIA
    }
    agg_spec.update(
        {
            f"{key}_support_3_4": (f"{key}_1_4", lambda s: (s >= 3).mean())
            for key, _ in SEED_CRITERIA
        }
    )
    grouped = (
        seeds_df.groupby(["seed_id", "seed_title", "target_severity_label"], as_index=False)
        .agg(
            **agg_spec,
            overall_mean_1_4=("overall_mean_1_4", "mean"),
            overall_support_3_4=("overall_mean_1_4", lambda s: (s >= 3).mean()),
            target_match_share=("severity_matches_target", "mean"),
            n_comments=("comment", lambda s: int(s.str.len().gt(0).sum())),
        )
        .sort_values(["seed_id"])
    )

    modal_labels = []
    for seed_id in grouped["seed_id"]:
        seed_subset = seeds_df.loc[seeds_df["seed_id"] == seed_id, "severity_rating_label"]
        modal_label = sorted(Counter(seed_subset).items(), key=lambda pair: (-pair[1], pair[0]))[0][0]
        modal_labels.append(modal_label)
    grouped["modal_severity_label"] = modal_labels
    grouped["severity_follow_up_flag"] = grouped["target_match_share"] < 0.5
    return grouped.sort_values(["overall_mean_1_4", "overall_support_3_4", "seed_id"]).reset_index(drop=True)


def build_comment_log(experts_df: pd.DataFrame, seeds_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, expert in experts_df.iterrows():
        if expert["tool_comment"]:
            rows.append(
                {
                    "comment_scope": "tool_overall",
                    "expert_id": expert["expert_id"],
                    "target_id": "tool",
                    "comment_text": expert["tool_comment"],
                }
            )
        if expert["seed_overall_comment"]:
            rows.append(
                {
                    "comment_scope": "seedbook_overall",
                    "expert_id": expert["expert_id"],
                    "target_id": "seedbook",
                    "comment_text": expert["seed_overall_comment"],
                }
            )
    for _, seed_row in seeds_df.iterrows():
        if seed_row["comment"]:
            rows.append(
                {
                    "comment_scope": "seed_specific",
                    "expert_id": seed_row["expert_id"],
                    "target_id": f"S{int(seed_row['seed_id']):02d}",
                    "comment_text": seed_row["comment"],
                }
            )
    return pd.DataFrame(rows)


def prepare_table_s4(item_summary: pd.DataFrame) -> pd.DataFrame:
    table = item_summary.copy()
    table["support_3_4_pct"] = (table["overall_support_3_4"] * 100).round(1)
    return table[
        [
            "item_code",
            "item_label",
            "relevance_mean_1_4",
            "clarity_mean_1_4",
            "need_mean_1_4",
            "overall_mean_1_4",
            "support_3_4_pct",
            "follow_up_flag",
        ]
    ].rename(
        columns={
            "item_code": "Kód",
            "item_label": "Položka",
            "relevance_mean_1_4": "Relevantnosť M",
            "clarity_mean_1_4": "Zrozumiteľnosť M",
            "need_mean_1_4": "Potrebnosť M",
            "overall_mean_1_4": "Celkový priemer M",
            "support_3_4_pct": "Podpora 3-4 (%)",
            "follow_up_flag": "Potrebný follow-up",
        }
    ).round(
        {
            "Relevantnosť M": 2,
            "Zrozumiteľnosť M": 2,
            "Potrebnosť M": 2,
            "Celkový priemer M": 2,
        }
    )


def prepare_table_s5(seed_summary: pd.DataFrame) -> pd.DataFrame:
    table = seed_summary.copy()
    table["target_match_pct"] = (table["target_match_share"] * 100).round(1)
    return table[
        [
            "seed_id",
            "seed_title",
            "overall_mean_1_4",
            "information_sufficiency_mean_1_4",
            "anti_stereotypy_mean_1_4",
            "research_usefulness_mean_1_4",
            "target_severity_label",
            "modal_severity_label",
            "target_match_pct",
        ]
    ].rename(
        columns={
            "seed_id": "Seed",
            "seed_title": "Scenár",
            "overall_mean_1_4": "Celkový priemer M",
            "information_sufficiency_mean_1_4": "Dostatok informácií M",
            "anti_stereotypy_mean_1_4": "Nestereotypnosť M",
            "research_usefulness_mean_1_4": "Vhodnosť pre výskum M",
            "target_severity_label": "Cieľová závažnosť",
            "modal_severity_label": "Modálna expert. závažnosť",
            "target_match_pct": "Zhoda so závažnosťou (%)",
        }
    ).round(
        {
            "Celkový priemer M": 2,
            "Dostatok informácií M": 2,
            "Nes stereotypnosť M": 2,
            "Vhodnosť pre výskum M": 2,
        }
    )


def write_combined_preview(table_s4: pd.DataFrame, table_s5: pd.DataFrame) -> None:
    html = f"""<!doctype html>
<html lang="sk">
<head>
  <meta charset="utf-8">
  <title>Expert Review Preview</title>
  <style>
    body {{
      margin: 0;
      background: #fbf7f2;
      color: #202529;
      font-family: Georgia, 'Times New Roman', serif;
    }}
    .wrap {{
      max-width: 1280px;
      margin: 0 auto;
      padding: 32px 48px 56px;
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 34px;
    }}
    p {{
      color: #5e646b;
      line-height: 1.5;
      font-size: 17px;
    }}
    h2 {{
      margin-top: 36px;
      font-size: 24px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      margin-top: 12px;
      background: #fffdfa;
      font-size: 14px;
    }}
    thead th {{
      text-align: left;
      padding: 10px 12px;
      background: #efeae2;
      border-top: 2px solid #525a61;
      border-bottom: 1px solid #b2b8bf;
    }}
    tbody td {{
      padding: 9px 12px;
      border-bottom: 1px solid #e1ddd6;
      vertical-align: top;
    }}
    tbody tr:nth-child(even) {{
      background: #faf6f0;
    }}
    img {{
      width: 100%;
      height: auto;
      border: 1px solid #d9d1c5;
      background: #fffdfa;
      margin-top: 12px;
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>Preview predbežnej expertnej obsahovej kontroly</h1>
    <p>
      Tento preview balík sumarizuje pilotný expert review pass pre položky ratingového nástroja a seed scenáre.
      Detailné CSV ostávajú v <code>tables/</code> a obrázky v <code>figures/</code>.
    </p>

    <h2>Tabuľka S4: Položky ratingového nástroja</h2>
    {table_s4.to_html(index=False, border=0)}
    <img src="../../figures/{FIGURE_ITEMS_PATH.name}" alt="Heatmap expert review položiek">

    <h2>Tabuľka S5: Seed scenáre</h2>
    {table_s5.to_html(index=False, border=0)}
    <img src="../../figures/{FIGURE_SEEDS_PATH.name}" alt="Heatmap expert review seedov">
  </div>
</body>
</html>
"""
    HTML_PREVIEW_PATH.write_text(html, encoding="utf-8")


def main() -> None:
    ensure_dirs()
    raw_df = read_named_columns_from_xlsx(RAW_EXPORT_PATH)

    experts_df = build_validation_experts(raw_df)
    items_df = build_item_review(raw_df)
    seeds_df = build_seed_review(raw_df)

    experts_df.to_csv(VALIDATION_EXPERTS_PATH, index=False)
    items_df.to_csv(ITEMS_CLEAN_PATH, index=False)
    seeds_df.to_csv(SEEDS_CLEAN_PATH, index=False)

    comment_log = build_comment_log(experts_df, seeds_df)
    panel_summary, panel_breakdown = summarize_panel(experts_df, items_df, seeds_df, comment_log)
    item_summary = summarize_items(items_df)
    seed_summary = summarize_seeds(seeds_df)

    panel_summary.to_csv(PANEL_SUMMARY_PATH, index=False)
    panel_breakdown.to_csv(PANEL_BREAKDOWN_PATH, index=False)
    item_summary.to_csv(ITEM_SUMMARY_PATH, index=False)
    seed_summary.to_csv(SEED_SUMMARY_PATH, index=False)
    comment_log.to_csv(COMMENT_LOG_PATH, index=False)

    table_s4 = prepare_table_s4(item_summary)
    table_s5 = prepare_table_s5(seed_summary)
    table_s4.to_csv(TABLE_ITEMS_PATH, index=False)
    table_s5.to_csv(TABLE_SEEDS_PATH, index=False)

    draw_heatmap(
        data=item_summary.assign(
            label=lambda df: df["item_code"] + " — " + df["item_label"]
        ),
        row_label_col="label",
        value_cols=[
            "relevance_mean_1_4",
            "clarity_mean_1_4",
            "need_mean_1_4",
            "overall_mean_1_4",
        ],
        display_cols=["Relevantnosť", "Zrozumiteľnosť", "Potrebnosť", "Celkový priemer"],
        title="Obrázok S3. Predbežná expertná kontrola položiek ratingového nástroja",
        subtitle="Priemerné skóre na škále 1-4; položky sú zoradené od najnižšieho celkového priemeru.",
        output_path=FIGURE_ITEMS_PATH,
    )
    draw_heatmap(
        data=seed_summary.assign(label=lambda df: "S" + df["seed_id"].map(lambda x: f"{int(x):02d}") + " — " + df["seed_title"]),
        row_label_col="label",
        value_cols=[
            "ambulatory_profile_mean_1_4",
            "context_realism_mean_1_4",
            "symptom_plausibility_mean_1_4",
            "clarity_mean_1_4",
            "information_sufficiency_mean_1_4",
            "anti_stereotypy_mean_1_4",
            "research_usefulness_mean_1_4",
            "overall_mean_1_4",
        ],
        display_cols=[
            "Ambulantný profil",
            "Kontext",
            "Symptómy",
            "Jasnosť",
            "Dostatok informácií",
            "Nestereotypnosť",
            "Vhodnosť pre výskum",
            "Celkový priemer",
        ],
        title="Obrázok S4. Predbežná expertná kontrola seed scenárov",
        subtitle="Priemerné skóre na škále 1-4; scenáre sú zoradené od najnižšieho celkového priemeru.",
        output_path=FIGURE_SEEDS_PATH,
    )

    write_html_table(
        table_s4,
        "Tabuľka S4. Predbežná expertná kontrola položiek ratingového nástroja",
        "Hodnotenia vychádzajú zo šiestich expertov. Stĺpec podpory udáva podiel ratingov 3-4 v percentách.",
        HTML_ITEMS_PATH,
    )
    write_html_table(
        table_s5,
        "Tabuľka S5. Predbežná expertná kontrola seed scenárov",
        "Hodnotenia vychádzajú zo šiestich expertov. Pri severity je uvedená cieľová a modálna expertne vnímaná závažnosť.",
        HTML_SEEDS_PATH,
    )
    write_combined_preview(table_s4, table_s5)

    print(f"Wrote {VALIDATION_EXPERTS_PATH.relative_to(REPO_ROOT)}")
    print(f"Wrote {ITEMS_CLEAN_PATH.relative_to(REPO_ROOT)}")
    print(f"Wrote {SEEDS_CLEAN_PATH.relative_to(REPO_ROOT)}")
    print(f"Wrote {TABLE_ITEMS_PATH.relative_to(REPO_ROOT)}")
    print(f"Wrote {TABLE_SEEDS_PATH.relative_to(REPO_ROOT)}")
    print(f"Wrote {FIGURE_ITEMS_PATH.relative_to(REPO_ROOT)}")
    print(f"Wrote {FIGURE_SEEDS_PATH.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
